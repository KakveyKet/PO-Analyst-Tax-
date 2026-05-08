import streamlit as st
import pdfplumber
import pandas as pd
import re
import io
import os
import openpyxl
import sqlite3
from copy import copy
from datetime import datetime, date
from openpyxl.styles import Font

# Import settings and functions from the styles.py file
from styles import (
    TEMPLATE_PATH, MARKET_MAPPING, THIN_BORDER, THICK_BORDER, 
    GREEN_DATA_FONT, FONT_HEADER_DEFAULT, FONT_HEADER_RED, 
    HEADER_ALIGNMENT, DATA_ALIGNMENT, HEADER_FILL_NEUTRAL, 
    NONE_TEMPLATE_HEADER_CONFIG, format_factory_name, format_ak_column,
    clean_none_template_val
)

# --- 1. THE EXTRACTION LOGIC ---
def parse_purchase_order(file_object):
    po_details = {
        "PO Number": "Not Found", "Date": "Not Found", "Remark Size": "Not Found", 
        "Style Name": "Not Found", "Age Group": "Not Found", "Gender": "Not Found", 
        "Garment Type": "Not Found", "Product Group": "Apparel", 
        "Factory Line": "Trax Apparel (Cambodia) Co.,Ltd.", "Product Division": "APP", 
        "Leather Logo": "No", "COO": "MADE IN CAMBODIA", "Season": "Not Found"
    }
    items_data = []

    try:
        with pdfplumber.open(file_object) as pdf:
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text(x_tolerance=1.2, y_tolerance=3)
                if text:
                    full_text += text + "\n"

            po_match = re.search(r"K[A-Z]{2,5}\d{6,}", full_text)
            if po_match: 
                po_details["PO Number"] = po_match.group(0)
            else:
                po_details["PO Number"] = file_object.name.replace(".pdf", "").replace(".PDF", "")

            size_match = re.search(r"SIZE\s*PAGE\s*([^\n]+)", full_text, re.IGNORECASE)
            if size_match: 
                po_details["Remark Size"] = size_match.group(1).strip()

            garment_patterns = {
                "Style Name": r"Style\s*Name[\s:;]*(.*?)(?=;|Age|Gender|Garment|Product|COO)",
                "Age Group": r"Age\s*Group[\s:;]*(.*?)(?=;|Gender|Garment|Product|COO)",
                "Gender": r"Gender[\s:;]*(.*?)(?=;|Garment|Product|COO)",
                "Garment Type": r"Garment\s*Type[\s:;]*(.*?)(?=;|Product|Factory|COO)",
                "Product Group": r"Product\s*Group[/\w\s]*[\s:;]*(.*?)(?=;|Factory|product)",
                "Factory Line": r"Factory\s*Line\s*name[\s:;]*(.*?)(?=;|product|Made|COO)",
                "Product Division": r"product\s*division[\s:;]*(.*?)(?=;|Made|COO)",
                "Leather Logo": r"Made\s*by\s*leather[\s:;]*(.*?)(?=;|COO)",
                "COO": r"COO[\s:;]*(.*?)(?=;|\n|001A)",
                "Season": r"Season[\s:;]*([A-Z0-9]+)"
            }

            for key, pattern in garment_patterns.items():
                match = re.search(pattern, full_text, re.IGNORECASE | re.DOTALL)
                if match:
                    clean_val = match.group(1).replace('\n', ' ').strip()
                    clean_val = re.sub(r'\s+', ' ', clean_val)
                    clean_val = clean_val.rstrip(';/').strip()
                    po_details[key] = clean_val

            # --- GLOBAL FALLBACKS ---
            global_fallback_style = "Not Found"
            cladd_style = re.search(r"CLADD Additional Care Label\s*\n\s*([A-Z0-9]+)", full_text, re.IGNORECASE)
            if cladd_style:
                global_fallback_style = cladd_style.group(1).strip()
            else:
                any_item_with_style = re.search(r"Item\s*[:;]?\s*\d{8}(?:-[A-Za-z0-9]+)*-([A-Z0-9]{4,})", full_text, re.IGNORECASE)
                if any_item_with_style:
                    global_fallback_style = any_item_with_style.group(1).strip()

            global_fallback_market = "Other"
            market_matches = re.findall(r"Order[\s:;-]*([A-Za-z]+)", full_text, re.IGNORECASE)
            skip_words = ["QUANTITY", "FORM", "DATE", "NUMBER", "NO", "PCS", "TRANSFER", "FOR", "MARKET", "GROUP", "TYPE"]
            
            if market_matches:
                for match in market_matches:
                    clean_match = match.strip().upper()
                    if clean_match not in skip_words:
                        global_fallback_market = MARKET_MAPPING.get(clean_match, clean_match)
                        break

            global_fallback_item_code = "Not Found"
            any_item_code = re.search(r"Item\s*[:;]?\s*(\d{8})", full_text, re.IGNORECASE)
            if any_item_code:
                global_fallback_item_code = any_item_code.group(1).strip()

            # --- MULTI-GROUP LINE-BY-LINE EXTRACTION ---
            current_market = global_fallback_market
            current_item_id = global_fallback_style
            current_item_code = global_fallback_item_code

            lines = full_text.split('\n')
            for i, line in enumerate(lines):
                # 1. Look for explicit market markers (e.g. "Order INDO;")
                market_match = re.search(r"Order[\s:;-]*([A-Za-z]+)", line, re.IGNORECASE)
                if market_match:
                    clean_match = market_match.group(1).strip().upper()
                    if clean_match not in skip_words:
                        current_market = MARKET_MAPPING.get(clean_match, clean_match)

                # 2. Look for item strings (e.g. 62759440-INDO-S2662...)
                item_match = re.search(r"Item\s*[:;]?\s*(\d{8}(?:-[A-Za-z0-9]+)*)", line, re.IGNORECASE)
                if item_match:
                    full_item_str = item_match.group(1)
                    parts = full_item_str.split('-')
                    current_item_code = parts[0]
                    if len(parts) >= 3:
                        mid_market = parts[1].strip().upper()
                        if mid_market not in skip_words:
                            current_market = MARKET_MAPPING.get(mid_market, mid_market)
                        current_item_id = "-".join(parts[2:]).strip()
                    elif len(parts) == 2:
                        val = parts[1].upper()
                        # Verify if the 2nd part is a market code or a style code
                        if val in MARKET_MAPPING or val in ["OTHER", "OTHERS", "INDONESIA", "KOREA", "CHINA", "JAPAN"]:
                            current_market = MARKET_MAPPING.get(val, val)
                        else:
                            current_item_id = parts[1].strip()

                # 3. Process the actual row of Data
                if "PCS" in line:
                    p = line.split()
                    if "PCS" in p:
                        if len(p) > 2 and (p[1] == "NO" or p[2] == "NO"): continue
                        pcs_idx = p.index("PCS")
                        try:
                            items_data.append({
                                "PO Number": po_details["PO Number"], 
                                "Item Code": current_item_code,
                                "Item ID": current_item_id,
                                "Order": current_market,
                                "Material": p[0], "Color": p[1], "Size": p[2], "Sub-Code": p[3],
                                "Qty": float(p[pcs_idx - 1]), "Unit": "PCS",
                                "Price": float(p[pcs_idx + 1]), "Amount": float(p[-2]), "Delivery": p[-1]
                            })
                        except (ValueError, IndexError): continue
                            
    except Exception as e:
        st.error(f"Error reading PDF: {e}")
        return None, None
    return po_details, pd.DataFrame(items_data)

# --- 2. THE EXACT-STYLE TEMPLATE POPULATOR ---
def populate_existing_template(template_path, df_merged):
    if df_merged.empty: return None
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active
    
    header_row = 19 
    for r in range(1, 50): 
        if "*PO# number" in str(sheet.cell(row=r, column=1).value):
            header_row = r
            break
    start_row = header_row + 1 

    max_template_col = 37 
    for c in range(100, 0, -1):
        if sheet.cell(row=header_row, column=c).value is not None and str(sheet.cell(row=header_row, column=c).value).strip() != "":
            max_template_col = c
            break

    for r in range(1, start_row + 1):
        ak_cell = sheet.cell(row=r, column=37)
        if ak_cell.value:
            ak_cell.value = format_ak_column(ak_cell.value)
            ak_cell.font = Font(name="Calibri", size=11)

    for index, row in df_merged.iterrows():
        current_row = start_row + index
        
        mapping = {
            1: row["*PO# number"], 2: row["*Style#"], 3: row["*Article#"], 4: row["*Product group / type "],
            5: row["*Style name"], 6: row["*COO"], 7: row["*Brand"], 
            8: row["*Factory line name "], 10: row["Leather logo (optional)"], 11: row["*Age group"], 
            12: row["*Size page"], 13: row["*Garment type"], 14: row["*Gender"], 15: row["*Product division "],
            19: row["*Sourcing size"], 22: row["*Order Quantity"], 23: row["*Order Market "], 
            24: row["*Order Type"], 25: row["*Season"]
        }

        for col_num in range(1, max_template_col + 1):
            target_cell = sheet.cell(row=current_row, column=col_num)
            
            if current_row != start_row: 
                template_cell = sheet.cell(row=start_row, column=col_num)
                if template_cell.has_style:
                    target_cell.fill = copy(template_cell.fill)
                    target_cell.number_format = copy(template_cell.number_format)
                    target_cell.protection = copy(template_cell.protection)
                    target_cell.alignment = copy(template_cell.alignment)
                    target_cell.font = copy(template_cell.font)
                if col_num not in mapping:
                    target_cell.value = template_cell.value

            if col_num in mapping:
                target_cell.value = mapping[col_num]
            
            target_cell.border = THIN_BORDER
            if col_num == 22: 
                target_cell.number_format = '0.00'
            if col_num == 37 and target_cell.value:
                target_cell.value = format_ak_column(target_cell.value)
                target_cell.font = Font(name="Calibri", size=11)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

# --- NONE TEMPLATE EXCEL GENERATOR ---
def generate_styled_none_template(df):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Order Form"
    
    for start_c, end_c, text in NONE_TEMPLATE_HEADER_CONFIG:
        sheet.merge_cells(start_row=1, start_column=start_c, end_row=1, end_column=end_c)
        main_cell = sheet.cell(row=1, column=start_c)
        main_cell.value = text
        current_font = FONT_HEADER_RED if text == "Chinese Customize Size - 2nd line size" else FONT_HEADER_DEFAULT
        for c in range(start_c, end_c + 1):
            cell = sheet.cell(row=1, column=c)
            cell.fill = HEADER_FILL_NEUTRAL
            cell.font = current_font
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THICK_BORDER
                
    for r_idx, row in enumerate(df.to_dict('records'), 2):
        for start_c, end_c, text_h in NONE_TEMPLATE_HEADER_CONFIG:
            if start_c != end_c:
                sheet.merge_cells(start_row=r_idx, start_column=start_c, end_row=r_idx, end_column=end_c)
            target_cell = sheet.cell(row=r_idx, column=start_c)
            if text_h == "PO":
                target_cell.value = clean_none_template_val(row.get("PO Number", ""))
            elif text_h == "customer ordered size":
                full_size = str(row.get("Sourcing Size", ""))
                size_parts = full_size.split('/')
                target_cell.value = size_parts[1] if len(size_parts) > 1 else full_size
            elif text_h == "Sourcing Size":
                full_size = str(row.get("Sourcing Size", ""))
                size_parts = full_size.split('/')
                target_cell.value = size_parts[0] if len(size_parts) > 1 else full_size
            elif text_h == "Tecn. Notation":
                target_cell.value = "B3"
            elif text_h == "size run ID":
                target_cell.value = "2X"
            elif text_h == "Chinese Customize Size - 2nd line size":
                target_cell.value = "NO"
            elif text_h == "Quantity":
                target_cell.value = float(row.get("Order Quantity", 0))
                target_cell.number_format = '0.00'
            elif text_h == "Age group":
                target_cell.value = clean_none_template_val(row.get("Age Group", ""))
            elif text_h == "Gender":
                target_cell.value = clean_none_template_val(row.get("Gender", ""))
            elif text_h == "Size Spec / Pattern":
                target_cell.value = "GLOBAL SIZE"
            elif text_h == "Size page":
                target_cell.value = clean_none_template_val(row.get("Size Page", ""))
            elif text_h == "Top/Bottom":
                target_cell.value = clean_none_template_val(row.get("Garment Type", "")).title()
            elif text_h == "Product division":
                target_cell.value = clean_none_template_val(row.get("Product Division", ""))
            elif text_h == "Season":
                target_cell.value = clean_none_template_val(row.get("Season", ""))
            elif text_h == "Style":
                style_val = clean_none_template_val(row.get("Style ID", ""))
                if not style_val: style_val = clean_none_template_val(row.get("Style Name", ""))
                target_cell.value = style_val
            elif text_h == "Code item":
                target_cell.value = clean_none_template_val(row.get("Item Code", ""))
            for c in range(start_c, end_c + 1):
                cell = sheet.cell(row=r_idx, column=c)
                cell.font = GREEN_DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = DATA_ALIGNMENT
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

# --- 3. THE UI & NAVIGATION ---
st.set_page_config(page_title="PO Data Extractor", page_icon=":material/description:", layout="wide")

# --- HIGH-VISIBILITY VUE-JS ROUTER CSS ---
st.markdown("""
    <style>
    /* 1. Hides the radio circles entirely */
    [data-testid="stSidebar"] div[role="radiogroup"] > label > div:first-child {
        display: none !important;
    }
    
    /* 2. Base styling for the Route Buttons */
    [data-testid="stSidebar"] div[role="radiogroup"] label {
        padding: 14px 22px !important;
        border-radius: 0px 30px 30px 0px !important; 
        background-color: transparent !important;
        border: none !important;
        margin-bottom: 8px !important;
        cursor: pointer !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        display: flex !important;
        align-items: center !important;
        width: 100% !important;
        margin-left: -20px !important; /* Forces left indicator to the very edge */
        border-left: 6px solid transparent !important;
    }
    
    /* 3. Inactive text color and font size */
    [data-testid="stSidebar"] div[role="radiogroup"] label div p {
        color: #606266 !important;
        font-size: 16px !important;
        font-weight: 500 !important;
    }
    
    /* 4. Hover State - Subtle light gray */
    [data-testid="stSidebar"] div[role="radiogroup"] label:hover {
        background-color: #F5F7FA !important;
        transform: translateX(4px) !important;
    }
    
    /* 5. ACTIVE / SELECTED STATE (The "Standing On" highlight) */
    /* Target the container of the checked radio */
    [data-testid="stSidebar"] div[role="radiogroup"] [data-checked="true"] label {
        background-color: #ECF5FF !important; /* Soft active blue background */
        border-left: 6px solid #409EFF !important; /* THE VUE-JS PRIMARY BLUE INDICATOR */
        padding-left: 22px !important;
        box-shadow: 0 4px 10px rgba(64, 158, 255, 0.15) !important;
    }
    
    /* 6. Force text inside ACTIVE label to Primary Blue and BOLD */
    [data-testid="stSidebar"] div[role="radiogroup"] [data-checked="true"] label div p {
        color: #409EFF !important;
        font-weight: 700 !important;
    }
    
    /* 7. Force Material Icons inside ACTIVE label to Primary Blue */
    [data-testid="stSidebar"] div[role="radiogroup"] [data-checked="true"] label span {
        color: #409EFF !important;
    }
    </style>
    """, unsafe_allow_html=True)

if "preview_df" not in st.session_state:
    st.session_state.preview_df = None
if "report_name_default" not in st.session_state:
    st.session_state.report_name_default = "Extracted_PO_Data"
if "none_preview_df" not in st.session_state:
    st.session_state.none_preview_df = None
if "none_report_name_default" not in st.session_state:
    st.session_state.none_report_name_default = "None_Template_PO_Data"

st.sidebar.title("App Navigation")
menu = st.sidebar.radio("Navigate Systems:", [
    ":material/download: 1. Data Extraction", 
    ":material/settings: 2. Data Processing & Export", 
    ":material/note_add: 3. None Template Data Extraction",
    ":material/build: 4. None Template Data Process",
    ":material/bar_chart: 5. Reports"
])

# =====================================================================
# 1. DATA EXTRACTION
# =====================================================================
if menu == ":material/download: 1. Data Extraction":
    st.title(":material/download: Data Extraction")
    st.markdown("Upload your PO PDFs for standard template processing.")
    uploaded_pdfs = st.file_uploader("Upload PO PDFs", type=["pdf"], accept_multiple_files=True)
    if uploaded_pdfs:
        all_po_details, all_items_data = [], []
        with st.spinner('Analyzing Purchase Orders...'):
            for file in uploaded_pdfs:
                details, df_items = parse_purchase_order(file)
                if details:
                    all_po_details.append(details)
                    if not df_items.empty: all_items_data.append(df_items)
        if all_po_details:
            master_details_df = pd.DataFrame(all_po_details)
            master_items_df = pd.concat(all_items_data, ignore_index=True) if all_items_data else pd.DataFrame()
            preview_rows = []
            
            df_merged_raw = pd.merge(master_items_df, master_details_df, on="PO Number", how="left")
            for _, row in df_merged_raw.iterrows():
                age_group_val = str(row["Age Group"]).capitalize() if pd.notna(row["Age Group"]) and str(row["Age Group"]).upper() != "NOT FOUND" else ""
                
                raw_market = str(row["Order"]).strip()
                market_found = "Other" if not raw_market or raw_market.upper() in ["NOT FOUND", "", "ORDER"] else raw_market
                
                preview_rows.append({
                    "*PO# number": row["PO Number"], "Item Code": row["Item Code"], "*Style#": row["Item ID"],
                    "*Article#": row["Sub-Code"], "*Product group / type ": row["Product Group"], "*Style name": row["Style Name"],
                    "*COO": "MADE IN CAMBODIA", "*Brand": "Adidas", "*Factory line name ": format_factory_name(row["Factory Line"]),
                    "Leather logo (optional)": row["Leather Logo"], "*Age group": age_group_val, "*Size page": row["Remark Size"],
                    "*Garment type": row["Garment Type"], "*Gender": row["Gender"], "*Product division ": row["Product Division"],
                    "*Sourcing size": row["Size"], "*Order Quantity": row["Qty"], "*Order Market ": market_found,
                    "*Order Type": "BULK", "*Season": row["Season"]
                })
            st.session_state.preview_df = pd.DataFrame(preview_rows)
            st.success("Extraction Complete! Go to Tab 2 to Export.")
            st.markdown("---")
            st.subheader("Database Injection")
            report_name = st.text_input("Batch ID:", value=st.session_state.report_name_default)
            if st.button("Inject Data"):
                db_df = st.session_state.preview_df.copy()
                db_df.insert(0, "Uploaded_Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                db_df.insert(0, "Report_Name", report_name)
                try:
                    conn = sqlite3.connect('po_database.db')
                    db_df.to_sql('uploaded_reports', conn, if_exists='append', index=False)
                    conn.close()
                    st.success("Successfully injected records!")
                except Exception as e: st.error(f"Error: {e}")

# =====================================================================
# 2. DATA PROCESSING & EXPORT
# =====================================================================
elif menu == ":material/settings: 2. Data Processing & Export":
    st.title(":material/settings: Data Processing & Export")
    if st.session_state.preview_df is not None:
        st.markdown("### Sort Your Data")
        sort_c1, sort_c2 = st.columns(2)
        sort_column = sort_c1.selectbox("Column:", ["(No Sorting)"] + list(st.session_state.preview_df.columns))
        sort_order = sort_c2.radio("Order:", ["Ascending", "Descending"], horizontal=True)
        export_df = st.session_state.preview_df.copy()
        if sort_column != "(No Sorting)":
            export_df = export_df.sort_values(by=sort_column, ascending=(sort_order=="Ascending")).reset_index(drop=True)
        st.dataframe(export_df, use_container_width=True)
        if not os.path.exists(TEMPLATE_PATH): st.error("Template Missing!")
        else:
            if st.button("Generate Excel"):
                output = populate_existing_template(TEMPLATE_PATH, export_df)
                st.download_button("Download Template Excel", output, f"{st.session_state.report_name_default}.xlsx")
    else: st.info("No data extracted yet. Please use Tab 1 first.")

# =====================================================================
# 3. NONE TEMPLATE EXTRACTION
# =====================================================================
elif menu == ":material/note_add: 3. None Template Data Extraction":
    st.title(":material/note_add: None Template Data Extraction")
    uploaded_pdfs = st.file_uploader("Upload PO PDFs (None Template)", type=["pdf"], accept_multiple_files=True)
    if uploaded_pdfs:
        all_po_details, all_items_data = [], []
        with st.spinner('Analyzing...'):
            for file in uploaded_pdfs:
                details, df_items = parse_purchase_order(file)
                if details:
                    all_po_details.append(details)
                    if not df_items.empty: all_items_data.append(df_items)
        if all_po_details:
            master_details_df = pd.DataFrame(all_po_details)
            master_items_df = pd.concat(all_items_data, ignore_index=True) if all_items_data else pd.DataFrame()
            preview_rows = []
            
            df_merged_raw = pd.merge(master_items_df, master_details_df, on="PO Number", how="left")
            for _, row in df_merged_raw.iterrows():
                age_group_val = str(row["Age Group"]).capitalize() if pd.notna(row["Age Group"]) and str(row["Age Group"]).upper() != "NOT FOUND" else ""
                
                raw_market = str(row["Order"]).strip()
                market_found = "Other" if not raw_market or raw_market.upper() in ["NOT FOUND", "", "ORDER"] else raw_market
                
                preview_rows.append({
                    "PO Number": row["PO Number"], "Item Code": row["Item Code"], "Style ID": row["Item ID"],
                    "Style Name": row["Style Name"], "Article Code": row["Sub-Code"], "Product Group": row["Product Group"],
                    "Country of Origin": "MADE IN CAMBODIA", "Brand": "Adidas", "Factory Line": format_factory_name(row["Factory Line"]),
                    "Leather Logo": row["Leather Logo"], "Age Group": age_group_val, "Size Page": row["Remark Size"],
                    "Garment Type": row["Garment Type"], "Gender": row["Gender"], "Product Division": row["Product Division"],
                    "Sourcing Size": row["Size"], "Order Quantity": row["Qty"], "Order Market": market_found, "Order Type": "BULK", "Season": row["Season"]
                })
            st.session_state.none_preview_df = pd.DataFrame(preview_rows)
            st.success("Raw Extraction Complete! Go to Tab 4 to process.")
            if st.button("Save Raw Data"):
                db_df = st.session_state.none_preview_df.copy()
                db_df.insert(0, "Uploaded_Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                try:
                    conn = sqlite3.connect('po_database.db')
                    db_df.to_sql('uploaded_reports_none_template', conn, if_exists='append', index=False)
                    conn.close()
                    st.success("Saved to Raw DB!")
                except Exception as e: st.error(f"Error: {e}")

# =====================================================================
# 4. NONE TEMPLATE PROCESS
# =====================================================================
elif menu == ":material/build: 4. None Template Data Process":
    st.title(":material/build: None Template Data Process")
    if st.session_state.none_preview_df is not None:
        export_df = st.session_state.none_preview_df.copy()
        st.dataframe(export_df, use_container_width=True)
        if st.button("Export Styled Excel"):
            styled_excel = generate_styled_none_template(export_df)
            st.download_button("Download Styled Order Form", styled_excel, f"{st.session_state.none_report_name_default}.xlsx")
    else: st.info("No raw data extracted yet. Please use Tab 3 first.")

# =====================================================================
# 5. REPORTS
# =====================================================================
elif menu == ":material/bar_chart: 5. Reports":
    st.title(":material/bar_chart: Reports")
    if os.path.exists('po_database.db'):
        conn = sqlite3.connect('po_database.db')
        history_df = pd.read_sql("SELECT * FROM uploaded_reports ORDER BY Uploaded_Date DESC", conn)
        conn.close()
        st.dataframe(history_df, use_container_width=True)
    else: st.info("Database is empty.")