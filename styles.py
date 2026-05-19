from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.styles.colors import Color
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
import re
import openpyxl
from copy import copy
import io

# --- CONSTANTS ---
TEMPLATE_PATH = "POTemplate.xlsx"

# --- SMART MARKET EXPANSION DICTIONARY ---
MARKET_MAPPING = {
    "INDO": "INDONESIA",
    "KOR": "KOREA",
    "JAP": "JAPAN",
    "JPN": "JAPAN",
    "CHI": "CHINA",
    "CHN": "CHINA",
    "TAI": "TAIWAN",
    "TWN": "TAIWAN",
    "OTHER": "Other",
    "OTHERS": "Other"
}

# --- BORDERS ---
THIN_BORDER = Border(
    left=Side(style='thin', color='A5A5A5'), 
    right=Side(style='thin', color='A5A5A5'), 
    top=Side(style='thin', color='A5A5A5'), 
    bottom=Side(style='thin', color='A5A5A5')
)

THICK_BORDER = Border(
    left=Side(style='thin', color='000000'), 
    right=Side(style='thin', color='000000'), 
    top=Side(style='thin', color='000000'), 
    bottom=Side(style='thin', color='000000')
)

# --- FONTS ---
GREEN_DATA_FONT = Font(color="008608")
FONT_HEADER_DEFAULT = Font(bold=True, color="000000")
FONT_HEADER_RED = Font(bold=True, color="FF0000")

# --- ALIGNMENTS ---
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_ALIGNMENT = Alignment(horizontal='center', vertical='center')

# --- FILLS ---
HEADER_FILL_NEUTRAL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

# --- NONE TEMPLATE HEADER CONFIGURATION ---
NONE_TEMPLATE_HEADER_CONFIG = [
    (1, 1, "PO"),
    (2, 6, "customer ordered size"),
    (7, 11, "Sourcing Size"),
    (12, 16, "Tecn. Notation"),
    (17, 20, "size run ID"),
    (21, 24, "Chinese Customize Size - 2nd line size"),
    (25, 28, "Quantity"),
    (29, 33, "Age group"),
    (34, 38, "Gender"),
    (39, 43, "Size Spec / Pattern"),
    (44, 47, "Size page"),
    (48, 51, "Top/Bottom"),
    (52, 54, "Product division"),
    (55, 57, "Season"),
    (58, 60, "Style"),
    (61, 63, "Code item")
]

# --- UTILITY FUNCTIONS ---

def format_factory_name(name):
    if not name: return name
    name = str(name)
    name = re.sub(r'([a-zA-Z])\(', r'\1 (', name)
    name = re.sub(r'\)([a-zA-Z])', r') \1', name)
    name = re.sub(r'(Trax)(Apparel)', r'\1 \2', name, flags=re.IGNORECASE)
    return name.strip()

def format_ak_column(text_val):
    if not text_val: return text_val
    text_str = str(text_val)
    blue_color = Color(rgb="FF0070C0")
    red_color = Color(rgb="FFFF0000")
    blue_font = InlineFont(rFont="Calibri", sz=11, color=blue_color)
    red_font = InlineFont(rFont="Calibri", sz=11, color=red_color)

    if "between 50% to 74.9%" in text_str or "between 0% to 49.9%" in text_str:
        parts = re.split(r'(Goose down|Duck down)', text_str)
        rich_text_elements = []
        for part in parts:
            if not part: continue
            if part in ["Goose down", "Duck down"]:
                rich_text_elements.append(TextBlock(red_font, part))
            else:
                rich_text_elements.append(TextBlock(blue_font, part))
        return CellRichText(*rich_text_elements)

    if "Filling:" in text_str:
        return CellRichText(TextBlock(blue_font, text_str))

    return text_str

def clean_none_template_val(val):
    if not val: return ""
    v_str = str(val).strip()
    if v_str.upper() in ["NOT FOUND", "NAN", "NONE"]:
        return ""
    return v_str

# --- 2. THE EXACT-STYLE TEMPLATE POPULATOR ---
def populate_existing_template(template_path, df_merged):
    if df_merged.empty: return None
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active
    
    header_row = 19 
    for r in range(1, 50): 
        if "*PO# number" in str(sheet.cell(row=r, column=1).value):
            header_row = r
            break
    start_row = header_row + 1 

    max_template_col = 37 
    for c in range(100, 0, -1):
        if sheet.cell(row=header_row, column=c).value is not None and str(sheet.cell(row=header_row, column=c).value).strip() != "":
            max_template_col = c
            break

    for r in range(1, start_row + 1):
        ak_cell = sheet.cell(row=r, column=37)
        if ak_cell.value:
            ak_cell.value = format_ak_column(ak_cell.value)
            ak_cell.font = Font(name="Calibri", size=11)

    for index, row in df_merged.iterrows():
        current_row = start_row + index
        
        mapping = {
            1: row["*PO# number"], 2: row["*Style#"], 3: row["*Article#"], 4: row["*Product group / type "],
            5: row["*Style name"], 6: row["*COO"], 7: row["*Brand"], 
            8: row["*Factory line name "], 10: row["Leather logo (optional)"], 11: row["*Age group"], 
            12: row["*Size page"], 13: row["*Garment type"], 14: row["*Gender"], 15: row["*Product division "],
            19: row["*Sourcing size"], 22: row["*Order Quantity"], 23: row["*Order Market "], 
            24: row["*Order Type"], 25: row["*Season"]
        }

        for col_num in range(1, max_template_col + 1):
            target_cell = sheet.cell(row=current_row, column=col_num)
            
            if current_row != start_row: 
                template_cell = sheet.cell(row=start_row, column=col_num)
                if template_cell.has_style:
                    target_cell.fill = copy(template_cell.fill)
                    target_cell.number_format = copy(template_cell.number_format)
                    target_cell.protection = copy(template_cell.protection)
                    target_cell.alignment = copy(template_cell.alignment)
                    target_cell.font = copy(template_cell.font)
                if col_num not in mapping:
                    target_cell.value = template_cell.value

            if col_num in mapping:
                target_cell.value = mapping[col_num]
            
            target_cell.border = THIN_BORDER
            if col_num == 22: 
                target_cell.number_format = '0.00'
            if col_num == 37 and target_cell.value:
                target_cell.value = format_ak_column(target_cell.value)
                target_cell.font = Font(name="Calibri", size=11)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

# --- NONE TEMPLATE EXCEL GENERATOR ---
def generate_styled_none_template(df):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Order Form"
    
    for start_c, end_c, text in NONE_TEMPLATE_HEADER_CONFIG:
        sheet.merge_cells(start_row=1, start_column=start_c, end_row=1, end_column=end_c)
        main_cell = sheet.cell(row=1, column=start_c)
        main_cell.value = text
        current_font = FONT_HEADER_RED if text == "Chinese Customize Size - 2nd line size" else FONT_HEADER_DEFAULT
        for c in range(start_c, end_c + 1):
            cell = sheet.cell(row=1, column=c)
            cell.fill = HEADER_FILL_NEUTRAL
            cell.font = current_font
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THICK_BORDER
                
    for r_idx, row in enumerate(df.to_dict('records'), 2):
        for start_c, end_c, text_h in NONE_TEMPLATE_HEADER_CONFIG:
            if start_c != end_c:
                sheet.merge_cells(start_row=r_idx, start_column=start_c, end_row=r_idx, end_column=end_c)
            target_cell = sheet.cell(row=r_idx, column=start_c)
            if text_h == "PO":
                target_cell.value = clean_none_template_val(row.get("PO Number", ""))
            elif text_h == "customer ordered size":
                full_size = str(row.get("Sourcing Size", ""))
                size_parts = full_size.split('/')
                target_cell.value = size_parts[0] if len(size_parts) > 1 else full_size
            elif text_h == "Sourcing Size":
                full_size = str(row.get("Sourcing Size", ""))
                size_parts = full_size.split('/')
                target_cell.value = size_parts[1] if len(size_parts) > 1 else full_size
            elif text_h == "Tecn. Notation":
                # --- NOW FULLY DYNAMIC ---
                target_cell.value = clean_none_template_val(row.get("Technical Notation", "B3"))
            elif text_h == "size run ID":
                # --- NOW FULLY DYNAMIC ---
                target_cell.value = clean_none_template_val(row.get("Size Run ID", "2X"))
            elif text_h == "Chinese Customize Size - 2nd line size":
                target_cell.value = "NO"
            elif text_h == "Quantity":
                target_cell.value = float(row.get("Order Quantity", 0))
                target_cell.number_format = '0.00'
            elif text_h == "Age group":
                target_cell.value = clean_none_template_val(row.get("Age Group", ""))
            elif text_h == "Gender":
                target_cell.value = clean_none_template_val(row.get("Gender", ""))
            elif text_h == "Size Spec / Pattern":
                target_cell.value = "GLOBAL SIZE"
            elif text_h == "Size page":
                target_cell.value = clean_none_template_val(row.get("Size Page", ""))
            elif text_h == "Top/Bottom":
                target_cell.value = clean_none_template_val(row.get("Garment Type", "")).title()
            elif text_h == "Product division":
                target_cell.value = clean_none_template_val(row.get("Product Division", ""))
            elif text_h == "Season":
                target_cell.value = clean_none_template_val(row.get("Season", ""))
            elif text_h == "Style":
                style_val = clean_none_template_val(row.get("Style ID", ""))
                if not style_val: style_val = clean_none_template_val(row.get("Style Name", ""))
                target_cell.value = style_val
            elif text_h == "Code item":
                target_cell.value = clean_none_template_val(row.get("Item Code", ""))
            for c in range(start_c, end_c + 1):
                cell = sheet.cell(row=r_idx, column=c)
                cell.font = GREEN_DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = DATA_ALIGNMENT
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output